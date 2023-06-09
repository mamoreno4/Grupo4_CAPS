from random import uniform, seed, randint, gauss
from numpy.random import Generator, PCG64
import numpy as np
import pandas as pd
from scipy.stats import binom
import openpyxl
import itertools
import os
from datetime import datetime


#archivo con las clases y funciones del main
#Clase cuartel, considera los datos de cada cuartel
class Cuartel:

    def __init__(self, listav):
        self.cosechable=0
        self.factor=1
        self.id=listav[1]
        self.variedad=listav[5]
        self.precio=listav[6]
        self.cosecha_por_dia=dict()
        for i in range(1,366):
            self.cosecha_por_dia[i]=[]
        self.dia_optimo=listav[2]
        self.costo_hora=listav[8]
        self.transporte={"Machali":listav[9],"Chepica":listav[10],"Nancagua":listav[11]}
        self.despreciacion={1000:[0.8,0.8875,0.95,0.9875,1,0.9833,0.9333,0.85],3000:[0.6,0.775,0.9,0.975,1,0.9667,0.8667,0.7],6000:[0.1,0.4938,0.775,0.9438,1,0.9111,0.6444,0.2]}
    #agrega cosecha a la lista de cosechas del cuartel
    def agregar_cosecha(self,dia,cantidad_bodega):
        self.cosecha_por_dia[dia].append(cantidad_bodega)
        pass
    #calcula la cantidad cosechable del cuartel
    def set_cosechable(self,cosechable):
        self.cosechable=cosechable
        pass
    #calcula el factor de depreciacion del cuartel
    def gen_desp(self,dia,precio):
        desp=1
        for i in self.cosecha_por_dia[dia]:
            desp=self.despreciacion[precio][dia+4-int(self.dia_optimo)]
        return desp
    def __str__(self):
        return "-> {}".format(self.id)
    
    def __repr__(self):
        return "Cuartel {}".format(self.id)
    
#Clase Bodega, considera los datos de cada bodega
class Bodega:
    def __init__(self,ubicacion_tanques,nombre):
        self.id_tanque=0
        self.ubicacion=nombre
        self.tanques=[]
        self.tanques_fermentando=[]
        self.tanques_disponibles=[]
        for i in range(1,len(ubicacion_tanques.index)):
            Tanq=ubicacion_tanques.iloc[i]
            T=Tanque(Tanq[1],self.ubicacion,Tanq[0])
            self.agregar_tanque_disponible(T)
        pass
    #Devuelve el tanque con el id ingresado
    def devolver_tanque_id(self,id):
        for i in self.tanques:
            if i.id==id:
                return i
        return None
    #Agrega un tanque a la lista de tanques disponibles
    def agregar_tanque_disponible(self,tanque):
        self.tanques.append(tanque)
        self.tanques_disponibles.append(tanque)
        pass
    #Agregra un tanque a la lista de tanques fermentando
    def agregar_tanque_fermentando(self,tanque):
        self.tanques_disponibles.remove(tanque)
        self.tanques_fermentando.append(tanque)
        #print("Se agrega tanque {} a fermentar".format(tanque.id))
        pass
    #Devuelve la cantidad fermentada en el dia
    def revisar_tanques(self,dia):
        salidas=[]
        self.tanques_disponibles=[]
        self.tanques_fermentando=[]
        for i in self.tanques:
            if i.estado=="Disponible":
                self.tanques_disponibles.append(i)
            elif i.estado=="Fermentando":
                if i.dia_termino==dia:
                    salidas.append(i.vaciar_tanque())
                    self.tanques_disponibles.append(i)
                else:
                    self.tanques_fermentando.append(i)
        return salidas
    #Devuelve la capacidad de los tanques disponibles
    def tanques_capacidad(self,cap):
        disp=[]
        for i in self.tanques_disponibles:
            if i.capacidad==cap:
                disp.append(i)
        return disp
    
    def __repr__(self):
        return "Bodega {}".format(self.ubicacion)
    
#Clase Tanque, considera los datos de cada tanque
class Tanque:

    def __init__(self,capacidad,ubicacion,id):
        self.id = id
        self.capacidad=capacidad
        self.estado="Disponible"
        self.dia_inicial=0
        self.dia_promedio=0
        self.dia_termino=0
        self.cantidad_fermentado=0
        self.variedad_fermentando=""
        self.precio=0
        self.promedio_fermentacion = {'C': 12, 'CF': 13, 'CS': 12, 'G': 12, 'M': 13, 'S': 12, 'V': 13, 'SB': 13, 'Ch': 13}

        self.ubicacion=ubicacion
        self.generado=0
        pass
    #Pone el tanque en estado fermentando, y le pasa los datos de la fermentación a la función generar_dia
    def fermentar(self,cantidad,dia,variedad,precio,distr):
        if self.estado=="Disponible":
            self.estado="Fermentando"
            self.dia_inicial=dia
            self.cantidad_fermentado=cantidad
            self.variedad_fermentando=variedad
            self.dia_promedio=self.promedio_fermentacion[variedad]
            self.precio=precio
            self.dia_termino=dia+self.generar_dia(variedad,distr)
            #print("Se comienza a fermentar {} de variedad {} hasta el día {} en la bodega {}".format(cantidad,variedad,self.dia_termino,self.ubicacion))
            pass
        else:
            print("Error: tanque no Disponible")
            pass
    #Vacía el tanque y devuelve los datos de la fermentación
    def vaciar_tanque(self):
        dias_fermentando=self.dia_termino-self.dia_inicial
        fermentado=[self.cantidad_fermentado, self.variedad_fermentando, dias_fermentando, self.precio]
        self.cantidad_fermentado=0
        self.variedad_fermentando=""
        self.dia_inicial=0
        self.dia_termino=0
        self.estado="Disponible"
        self.generado=0
        return fermentado
    #Genera un dia aleatorio de fermentacion
    def generar_dia(self,variedad,distr):
        n=distr.loc[variedad][1]
        p=distr.loc[variedad][2]
        dia_generado=binom.rvs(n, p)
        self.generado=dia_generado
        return dia_generado
    
    def __repr__(self):
        return "Tanque {}".format(self.id)+" de la bodega {}".format(self.ubicacion)
    

#Resumen de la simulación
class Resumen:

    def __init__(self):
        self.dias=dict()
        self.sobras_cepas_bodega={'Machali':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Chepica':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Nancagua':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}}
        self.dias_generados=[]
        self.cosechado=0
        self.fermentado=0
        self.sobras=0
        self.seed=0
        self.sobras_cantidad_dia=[]
        self.porcentaje_tanque=[]
        self.total_cosechable=0
        self.dias_ocupado_tanques=[]
        self.fermentado_1000 = {'Machali':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Chepica':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Nancagua':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}}
        self.fermentado_3000 = {'Machali':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Chepica':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Nancagua':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}}
        self.fermentado_6000 = {'Machali':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Chepica':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Nancagua':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}}
        self.costo_transporte=0
        self.costo_trabajo=0
        self.sobras_1000 = {'Machali':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Chepica':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Nancagua':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}}
        self.sobras_3000 = {'Machali':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Chepica':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Nancagua':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}}
        self.sobras_6000 = {'Machali':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Chepica':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}, 'Nancagua':{'G':0, 'Ch':0, 'SB':0, 'C':0, 'CS':0, 'S':0, 'M':0, 'CF':0, 'V':0}}
        self.costo_dias=0
        self.costo_sobras=0
        self.ganancias=0
        self.dict_dia=[]
        self.largo=0
        self.time=0
        self.sim=0
        for i in range(200):
            self.dias[i]=[]
        pass
    #agrega un dia a la lista de dias
    def agregar_dia(self,dia):
        self.dias[dia]=[]
        pass
    #agrega cosecha a la lista de cosechas
    def agregar_cosechaT(self,cuarteles):
        for i in cuarteles:
            self.total_cosechable+=i.cosechable
    def agregar_cosecha(self,dia,cosecha):
        self.dias[dia].append("Se cosecho "+cosecha)
        pass

    def agregar_fermentado(self,dia,fermentado):
        self.dias[dia].append("Se fermento "+str(fermentado[0])+" de variedad "+str(fermentado[1])+" de precio "+str(fermentado[3])+" en "+ str(fermentado[2]) +" dias")
        pass

    def agregar_sobrante(self,dia,sobrante,precio):
        self.dias[dia].append("El dia"+ str(dia) +" sobro la cantidad de "+str(sobrante)+" de variedad "+str(precio))
        pass

    def agregar_tanque(self, dia, cantidad):
        self.dias[dia].append("El dia"+ dia +" se agrego a tanque la cantidad de "+ cantidad)
        pass
    #Genera los promedios de los datos de los dias
    def gen_promedio(self):
        Ganancias_COSTO = self.ganancias-self.costo_dias-self.costo_sobras-self.costo_transporte-self.costo_trabajo

        Promedios = [["Dias generados",],["Dias Ocupado tanques",],["Promedio de porcentaje llenado tanque",],["Promedio de sobrante por dia",],["Cosechado",self.cosechado],["Fermentado",self.fermentado],["Nombre(seed)",self.seed],["Sobrante",self.sobras],["Costo dias",self.costo_dias],["Costo sobras",self.costo_sobras],["Ganancias",self.ganancias],["Costo transporte",self.costo_transporte],["Costo trabajo",self.costo_trabajo],["Ganancia Bruta",Ganancias_COSTO]]
        datos=[self.dias_generados,self.dias_ocupado_tanques,self.porcentaje_tanque,self.sobras_cantidad_dia]
        p=0
        for x in datos:
            Promedio = 0
            for i in x:
                Promedio += i
            if Promedio != 0:
                Promedi = Promedio/len(x)
            Promedios[p].append(Promedi)
            p+=1
        return Promedios
    #Imprime los datos de los dias
    def imprimir_resumen(self):
        f_total_1000 = 0
        f_total_3000 = 0
        f_total_6000 = 0
        s_total_1000 = 0
        s_total_3000 = 0
        s_total_6000 = 0
        for i in self.fermentado_1000:
            for j in self.fermentado_1000[i]:
                f_total_1000 += self.fermentado_1000[i][j]
                print("EL total final fermentado en la bodega {} de cepa {} es {}".format(i,j,self.fermentado_1000[i][j]))
        for i in self.fermentado_3000:
            for j in self.fermentado_3000[i]:
                f_total_3000 += self.fermentado_3000[i][j]
                print("EL total final fermentado en la bodega {} de cepa {} es {}".format(i,j,self.fermentado_3000[i][j]))
        for i in self.fermentado_6000:
            for j in self.fermentado_6000[i]:
                f_total_6000 += self.fermentado_6000[i][j]
                print("EL total final fermentado en la bodega {} de cepa {} es {}".format(i,j,self.fermentado_6000[i][j]))
        
        for i in self.sobras_1000:
            for j in self.sobras_1000[i]:
                s_total_1000 += self.sobras_1000[i][j]
                print("EL total final de sobras en la bodega {} de cepa {} es {}".format(i,j,self.sobras_1000[i][j]))
        for i in self.sobras_3000:
            for j in self.sobras_3000[i]:
                s_total_3000 += self.sobras_3000[i][j]
                print("EL total final de sobras en la bodega {} de cepa {} es {}".format(i,j,self.sobras_3000[i][j]))
        for i in self.sobras_6000:
            for j in self.sobras_6000[i]:
                s_total_6000 += self.sobras_6000[i][j]
                print("EL total final de sobras en la bodega {} de cepa {} es {}".format(i,j,self.sobras_6000[i][j]))
        
        # Final de fermentado y sobras por precio
        print("El total final de fermentado con precio 1000 es {}".format(f_total_1000))
        print("El total final de fermentado con precio 3000 es {}".format(f_total_3000))
        print("El total final de fermentado con precio 6000 es {}".format(f_total_6000))
        print("El total final de sobras con precio 1000 es {}".format(s_total_1000))
        print("El total final de sobras con precio 3000 es {}".format(s_total_3000))
        print("El total final de sobras con precio 6000 es {}".format(s_total_6000))
#Crea el excel con los datos de los dias
def crear_excel(nombre_archivo,lista_resumenes):
    #Toma la fecha y hora actual para crear la carpeta
    now = datetime.now()
    current_time = now.strftime("%H,%M,%S")
    #Crea la carpeta con el nombre del archivo y la fecha y hora
    path = './soluciones_implementacion/'+nombre_archivo+"_"+current_time
    os.makedirs(path)
    os.chdir(path)
    #Crea el excel con los datos de los dias
    for i in lista_resumenes:
        df_cosecha = pd.DataFrame({'Dia': i.dict_dia[0][0], 'Cuartel': i.dict_dia[0][1], 'cosecha': i.dict_dia[0][2]})
        df_trabajadores = pd.DataFrame({'Dia': i.dict_dia[1][0], 'Cuartel': i.dict_dia[1][1], 'Valor': i.dict_dia[1][2]})
        df_fermentado = pd.DataFrame({'Dia': i.dict_dia[2][0], 'Cepa': i.dict_dia[2][1], 'Calidad': i.dict_dia[2][2], 'Result': i.dict_dia[2][3]})
    open("cosecha.csv","w")
    df_cosecha.to_csv("cosecha.csv", index=False, mode="a")

    open("trabajadores.csv","w")
    df_trabajadores.to_csv("trabajadores.csv", index=False, mode="a")
    open("fermentado.csv","w")
    df_fermentado.to_csv("fermentado.csv", index=False, mode="a")
    nombre=nombre_archivo+current_time+".xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws['A1'] = "Nombre(seed)-largo periodo"
    ws['B1'] = "Cosechado"
    ws['C1'] = "No cosechado"
    ws['D1'] = "Fermentado"
    ws['E1'] = "Sobrante"
    ws['F1'] = "Promedio de sobrante por dia"
    ws['G1'] = "Promedio de porcentaje llenado tanque"
    ws['H1'] = "Dias generados"
    ws['I1'] = "Dias Ocupado tanques"
    ws['J1'] = "Costos dias fermentando"
    ws['K1'] = "Costos perdidas"
    ws['L1'] = "Costos Transporte"
    ws['M1'] = "Costos Trabajadores"
    ws['N1'] = "Ganancias"
    ws['O1'] = "Ganancias Bruta"
    ws['P1'] = "Tiempo se simulacion"
    ws['Q1'] = "Cantidad gurobi"
    c=2
    for i in lista_resumenes:
        G=i.gen_promedio()
        ws['A'+str(c)] = str(G[6][1])+"-"+str(i.largo)
        ws['B'+str(c)] = G[4][1]
        ws['C'+str(c)] = i.total_cosechable-i.cosechado
        ws['D'+str(c)] = G[5][1]
        ws['E'+str(c)] = G[7][1]
        ws['F'+str(c)] = G[3][1]
        ws['G'+str(c)] = G[2][1]
        ws['H'+str(c)] = G[0][1]
        ws['I'+str(c)] = G[1][1]
        ws['J'+str(c)] = G[8][1]
        ws['K'+str(c)] = G[9][1]
        ws['N'+str(c)] = G[10][1]
        ws['L'+str(c)] = G[11][1]
        ws['M'+str(c)] = G[12][1]
        ws['O'+str(c)] = G[13][1]
        ws['P'+str(c)] = i.time
        ws['Q'+str(c)] = i.sim


        c+=1
    wb.save(nombre)
    wb.close()
#encuentra la mejor combinacion de tanques para un liquido
def encontrar_combinacion_liquido(liquido, tanques):
    # Inicializar variables
    mejor_combinacion = None
    diferencia_minima = float('inf')
    num_tanques = len(tanques)
    # Iterar sobre todas las combinaciones de tanques
    for r in range(1, num_tanques + 1):
        # Iterar sobre todas las combinaciones de tanques de tamaño r
        for combo in itertools.combinations(tanques, r):
            # Calcular la suma de las capacidades de los tanques en la combinación
            suma_capacidades = sum(capacidad for _, capacidad in combo)
            # Verificar si la combinación es válida
            if liquido >= suma_capacidades * 0.74 and liquido <= suma_capacidades * 0.96:
                # Calcular la diferencia entre la suma de las capacidades y el líquido
                diferencia = abs(suma_capacidades - liquido)
                # Verificar si la diferencia es menor a la diferencia mínima
                if diferencia < diferencia_minima:
                    diferencia_minima = diferencia
                    mejor_combinacion = combo
    return mejor_combinacion

# funcion para llenar los tanques version final

def llenar_tanques(liquido_total, tanques):
    # Ordenar los tanques por capacidad (de mayor a menor)
    tanques_ordenados = sorted(tanques, key=lambda x: x[1], reverse=True)
    
    # Inicializar una lista para almacenar la cantidad de líquido en cada tanque
    liquido_tanques = []
    tanques_usados = []
    
    # Recorrer los tanques ordenados
    for i, tanque in enumerate(tanques_ordenados):
        nombre_tanque, capacidad_tanque = tanque
        tanques_usados.append(tanque)
        # Calcular la cantidad de líquido a poner en el tanque actual
        min_liquido = float(0.75 * capacidad_tanque)  # 75% de la capacidad del tanque
        max_liquido = float(0.95 * capacidad_tanque)  # 95% de la capacidad del tanque
        T=tanques_ordenados.copy()
        liquido_min=0
        for i in tanques_usados:
            if i in T:
                T.remove(i)
        for i in T:
            liquido_min+=0.75*i[1]
        tankL=min(max_liquido,liquido_total-liquido_min)

        cantidad_liquido = min(liquido_total, tankL)
        
        liquido_tanques.append((nombre_tanque, cantidad_liquido))
        liquido_total -= cantidad_liquido
        
        # Si ya no queda líquido por distribuir, se detiene el bucle
        if liquido_total == 0:
            break
    
    return liquido_tanques
# funcion para llenar los tanques version final
def comb_liquido(tanques, liquido):
    # Inicializar variables
    cantidad_liquido_tanques=None
    combinaciones = encontrar_combinacion_liquido(liquido, tanques)
    # Verificar si se encontró una combinación
    com_best=combinaciones
    # Verificar si se encontró una combinación
    if com_best:
        print(com_best)
        cantidad_liquido_tanques = llenar_tanques(liquido, com_best)
    return cantidad_liquido_tanques
#Pasar los tanques a un diccionario
def pasar_tanques_a_diario(tanques_ocupados):
    diccionario_datos_estanques = tanques_ocupados.to_dict(orient='records')
    dic_tanques_dia={}
    for i in range(1,150):
        di="dia_"+str(i)
        dic_tanques_dia[di]=[]
    for i in diccionario_datos_estanques:
        if i["Dia"] in dic_tanques_dia:
            dic_tanques_dia[i["Dia"]].append(i)
        else:
            dic_tanques_dia[i["Dia"]]=[i]
    return dic_tanques_dia
#funcion para revisar si los tanques estan ocupados o desocupados comparado con la planificacion
def revisar_input(dia_dict,tanques_ocupados,bodegas,dia):
    tanques_problemas=[]
    di="dia_"+str(dia)
    tanque_a_usar=dia_dict[di]
    for b in tanque_a_usar:
        id=int(b["Estanque"].split("_")[2])
        ub=b["Bodega"]
        for i in bodegas:
            if i.ubicacion==ub:
                for a in i.tanques_fermentando:
                    if a.id==id:
                        tanques_problemas.append([a,"tanque ocupado"])
    for b in tanques_ocupados:
        id=int(b.split("_")[2])
        ub=b.split("_")[1]
        ub=ub.capitalize()
        if di in tanques_ocupados[b]:
            for i in bodegas:
                if i.ubicacion==ub:
                    for a in i.tanques_disponibles:
                        if a.id==id:
                            tanques_problemas.append([a,"tanque disponible"])
    return tanques_problemas

     
#Pasar los tanques a un diccionario
def pasar_tanques_dict(BODEGAS,dia,largo):
    tanques_ocupados={}
    tanques_realidad={}
    for i in BODEGAS:
        bod=i.ubicacion
        for a in i.tanques_fermentando:
            id=str(a.id)
            nombre="estanque_"+bod+"_"+id
            nombre=nombre.lower()
            tanques_ocupados[nombre]={}
            tanques_realidad[nombre]={}
            diat=a.dia_termino
            for i in range(dia,min(min(max(a.dia_inicial+a.dia_promedio,dia+1),79),dia+largo)):
                tanques_ocupados[nombre]["dia_"+str(i)]=1
            for i in range(dia,a.dia_termino):
                tanques_realidad[nombre]["dia_"+str(i)]=1
                
    return tanques_ocupados,tanques_realidad
#Pasar los cuartel a un diccionario
def pasar_cuartel_dict(cuarteles):
    cosechar={}
    for i in cuarteles:
        nombre="cuartel_"+str(i.id)[:-2]
        cosechar[nombre]=max(round(i.cosechable/i.factor,2),0)
    return cosechar

#Creas las clases bases de los datos
def crear_clases():
    #Cargar datos
    Distribuciones = pd.read_excel('./../Distribuciones/dist.xlsx', index_col=0)
    cosecha_cuarteles=pd.read_excel('./../Distribuciones/estimado cosecha.xlsx')
    cosecha_cuarteles=cosecha_cuarteles.values.tolist()
    #Leer cuarteles
    Cuart = pd.read_excel('Datos Base Ordenados (Cosecha).xlsx')
    Cuartf = pd.read_excel('Datos base G4.xlsx')


    #PONER DATOS COSECHAS (ESTIMACION DE CADA CUARTEL)
    Los_Cuarteles = []
    #iterar por la cantidad de cuarteles
    for i in range(1,61):
        #crear cuartel
        CT = Cuartel(Cuart.iloc[i-1])
        CT.factor = Cuartf.iloc[i-1][6]
        CT.set_cosechable(cosecha_cuarteles[i-1][0])
        #agregar cuartel a la lista
        Los_Cuarteles.append(CT)
    #Leer bodegas
    Las_Bodegas = []
    Bodegas = pd.read_excel("Datos base G42.xlsx",sheet_name="Bodega Machali")
    #agregar datos de bodega a la lista, tambien agrega los tanques a las bodegas correspondientes
    b=Bodega(Bodegas,"Machali")
    Las_Bodegas.append(b)
    Bodegas = pd.read_excel("Datos base G42.xlsx",sheet_name="Bodega Chépica")
    b=Bodega(Bodegas,"Chepica")
    Las_Bodegas.append(b)
    Bodegas = pd.read_excel("Datos base G42.xlsx",sheet_name="Bodega Nancagua")
    b=Bodega(Bodegas,"Nancagua")
    Las_Bodegas.append(b)
    #lsita con el resumen de cada iteracon
    return Distribuciones,Los_Cuarteles,Las_Bodegas

def leer_gurobi(Los_Cuarteles,cosecha,trabajadores):
    diccionario_datos_cosecha = cosecha.to_dict(orient='records')
    diccionario_datos_trabajadores = trabajadores.to_dict(orient='records')
    for elemento in diccionario_datos_cosecha:
        for cuartel in Los_Cuarteles:
            C='cuartel_'+str(cuartel.id).split('.')[0]
            if C == elemento['Cuartel']:
                cuartel.agregar_cosecha(int(elemento['Dia'].split("_")[1]), [elemento["Bodega"],round(elemento['Valor'],2)])
    return diccionario_datos_trabajadores

def tanques_dia(dia,Nbodega,Ncepa,dict_diario,calidad):
    tanques_a_usar=[]
    if dia in dict_diario:
        for i in dict_diario[dia]:
            if i["Bodega"]==Nbodega and i["Cepa"]==Ncepa and int(i["Calidad"])==calidad:
                id_tanque=i["Estanque"].split("_")[2]
                tanques_a_usar.append([Nbodega,Ncepa,calidad,id_tanque])
    return tanques_a_usar
